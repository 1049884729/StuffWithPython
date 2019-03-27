#!/usr/bin/python
# -*- coding: UTF-8 -*-

import os
import os.path
import xlrd
import string

from openpyxl import Workbook

searchdir = "/study/FordExcels/excels"
mainFile = "aims.xlsx"


def main():
    searchFile = os.listdir(searchdir)
    listvalue=mainFileSearchKeyList(mainFile)
    listFiles=set()

    for f in searchFile:
        if (f.endswith("xlsx")):
            listFiles.add(os.path.join(searchdir,f))
    notConSet=set()
    for f in listFiles:
        tempSet=searchFileAndKey(listvalue,f)
        notConSet= notConSet|tempSet
    writelistData=readMainNotItem(notConSet,mainFile)
    writeExcel(writelistData)
def readMainNotItem(notConSet,file):
    """
    :return:
    """
    data = xlrd.open_workbook(file, "rb")
    sheet = data.sheet_by_name("Sheet1")
    listRow=list()
    print(notConSet)

    cols = sheet.ncols
    for i in notConSet:
        listColumen = list()
        for col in range(0,cols):
           rowData = sheet.cell(i, col).value
           listColumen.append(rowData)
        listRow.append(listColumen)
    return listRow


def searchFileAndKey(listvalue, file):
    data = xlrd.open_workbook(file, "rb")
    worksheet = data.sheet_by_index(0)
    rows = worksheet.nrows


    fileRowsData = list()
    for i in range(0, rows):
        rowData = worksheet.cell(i, 2).value
        fileRowsData.append(rowData)

    notContaintSet=set()
    for i,key in enumerate(listvalue):
        flag = False
        for row in fileRowsData:
            if (str(key) in str(row)):
                 flag=True
            if (flag):
                break

        if(not flag):
            print("Result:",flag,  key)
            notContaintSet.add(i)
    return notContaintSet




    # rowList =set()
    # for row in sheet.st
    # for key in listvalue:

def writeExcel(list):
    w = Workbook()  # 创建一个工作簿
    ws = w.create_sheet('Sheet')  # 创建一个工作表
    rows=len(list)
    for i in range(0, rows):  # 控制row
        columnList=list[i]
        print(columnList)
        ws.append(columnList)
    w.save('xqtest.xls')

def mainFileSearchKeyList(file):
    data = xlrd.open_workbook(file, "rb")
    sheet = data.sheet_by_name("Sheet1")

    listValue =list()
    for i in sheet.col_values(3):
        if (not i is "Issue Number"):
            listValue.append(i)
    return listValue


# mainFileSearchKeyList(mainFile)
main()